import os
import datetime
import openpyxl

def save_bdd_to_file(bdd_text, fmt, input_path, output_dir="output", rag_context=None):
    os.makedirs(output_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"{base_name}_BDD_{timestamp}.{fmt}")

    # Add RAG context as comment/header for text formats
    if rag_context and fmt in ["txt", "md", "feature"]:
        rag_comment = "\n\n# === RAG Context ===\n# " + "\n# ".join(rag_context.splitlines()) + "\n\n"
        bdd_text = rag_comment + bdd_text

    if fmt == "txt" or fmt == "md" or fmt == "feature":
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(bdd_text)
    elif fmt == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BDD Scenarios"
        for i, line in enumerate(bdd_text.splitlines(), start=1):
            ws.cell(row=i, column=1, value=line)
        # Save RAG context in a separate sheet if present
        if rag_context:
            ws2 = wb.create_sheet("RAG Context")
            for i, line in enumerate(rag_context.splitlines(), start=1):
                ws2.cell(row=i, column=1, value=line)
        wb.save(output_path)
    else:
        raise ValueError("Unsupported output format")

    return output_path
